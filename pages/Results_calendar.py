# 10_Results_Calendar.py
import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
from datetime import date, timedelta, datetime
import re
from pathlib import Path
from utils import *


tech_path = Path("data/technicals_all.parquet")

if tech_path.exists():
    df_tech = pd.read_parquet(tech_path)
else:
    df_tech = pd.DataFrame()
# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(layout="wide")
st.title("📅 Results Calendar")
st.caption("Upcoming quarterly result declarations — source: Trendlyne")

# ─── Fetch & parse ────────────────────────────────────────────────────────────

BASE_URL = "https://trendlyne.com/equity/calendar-v1/all/all/"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
    "Referer": "https://trendlyne.com/",
}


@st.cache_data(ttl=60 * 60 * 4, show_spinner=False)   # 4-hour cache
def fetch_results_calendar(start: str, end: str) -> pd.DataFrame:
    """
    Fetch the Trendlyne v1 calendar page and parse the HTML table.
    Filters server-side by posting the form; falls back to client-side
    date filtering if the POST doesn't apply the range.
    """
    params = {
        "start_date": start,
        "end_date":   end,
        "corporate_actions": "Results",
        "defaultStockgroup":  "nse/All/",
    }

    try:
        # First hit the homepage to get a session cookie (Trendlyne checks referer)
        session = requests.Session()
        session.headers.update(HEADERS)
        session.get("https://trendlyne.com/", timeout=10)

        resp = session.get(BASE_URL, params=params, timeout=20)
        resp.raise_for_status()
        html = resp.text
    except Exception as e:
        st.error(f"Could not reach Trendlyne: {e}")
        return pd.DataFrame()

    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if not table:
        st.warning("No calendar table found in the page response.")
        return pd.DataFrame()

    rows = []
    for tr in table.find_all("tr")[1:]:   # skip header
        cols = tr.find_all(["td", "th"])
        if len(cols) < 3:
            continue

        # ── Stock name + NSE symbol from the first cell's <a> href ────────────
        a_tag = cols[0].find("a", href=True)
        if not a_tag:
            continue
        company = a_tag.get_text(strip=True)
        href    = a_tag["href"]

        # URL pattern: /equity/corporate-actions/SYMBOL/ID/slug/
        sym_match = re.search(r"/equity/corporate-actions/([^/]+)/", href)
        symbol = sym_match.group(1).upper() if sym_match else ""

        # ── Date ──────────────────────────────────────────────────────────────
        date_raw = cols[1].get_text(strip=True)   # YYYY-MM-DD
        try:
            result_date = pd.to_datetime(date_raw).date()
        except Exception:
            continue

        # ── Event type ────────────────────────────────────────────────────────
        event = cols[2].get_text(strip=True)

        # ── Notes ─────────────────────────────────────────────────────────────
        notes = cols[3].get_text(strip=True) if len(cols) > 3 else ""

        rows.append({
            "symbol":      symbol,
            "company":     company,
            "result_date": result_date,
            "event":       event,
            "notes":       notes,
            "tl_url":      "https://trendlyne.com" + href,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # Client-side date filter (in case server didn't apply it)
    start_dt = pd.to_datetime(start).date()
    end_dt   = pd.to_datetime(end).date()
    df = df[
        df["result_date"].between(start_dt, end_dt) &
        df["event"].str.lower().str.contains("result|board meeting", na=False)
    ]

    return df.drop_duplicates(subset=["symbol", "result_date"]).reset_index(drop=True)


# ─── Load watchlist ───────────────────────────────────────────────────────────

WATCHLIST_FILE = Path("watchlist.txt")

@st.cache_data(ttl=3600)
def load_watchlist() -> set:
    if not WATCHLIST_FILE.exists():
        return set()
    symbols = set()
    for line in WATCHLIST_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("###"): continue
        for part in line.split(","):
            part = part.strip()
            if ":" in part: part = part.split(":", 1)[1]
            sym = re.sub(r"[^A-Z0-9&]", "", part.upper())
            if sym: symbols.add(sym)
    return symbols


# ─── Date controls ────────────────────────────────────────────────────────────

today    = date.today()
# Default: Monday → Sunday of current week
monday   = today - timedelta(days=today.weekday())
sunday   = monday + timedelta(days=6)

ctrl1, ctrl2, ctrl3 = st.columns([1, 1, 2])
with ctrl1:
    start_date = st.date_input("From", value=monday)
with ctrl2:
    end_date   = st.date_input("To",   value=sunday)
with ctrl3:
    show_mode  = st.radio(
        "Show",
        ["Watchlist stocks only", "All NSE stocks"],
        horizontal=True,
    )

if start_date > end_date:
    st.error("Start date must be before end date.")
    st.stop()

# ─── Fetch data ───────────────────────────────────────────────────────────────

with st.spinner("Fetching results calendar from Trendlyne…"):
    cal_df = fetch_results_calendar(
        start_date.strftime("%Y-%m-%d"),
        end_date.strftime("%Y-%m-%d"),
    )

if st.button("🔄 Refresh", use_container_width=False):
    st.cache_data.clear()
    st.rerun()

if cal_df.empty:
    st.info("No results found for the selected date range.")
    st.stop()

# ─── Enrich with watchlist & summary data ────────────────────────────────────

watchlist = load_watchlist()
cal_df["in_watchlist"] = cal_df["symbol"].isin(watchlist)

# Merge performance grade from summary if available
df_sh, df_fin, df_cf, df_insider, df_snap , df_brokerage, df_tech= load_all_data()
summary = st.session_state.get("summary", pd.DataFrame())
if summary.empty and not df_fin.empty:
    syms = tuple(cal_df["symbol"].unique().tolist())
    cats = ("Promoters", "FIIs", "DIIs", "Public")
    summary = build_summary_cached(
    df_sh,
    df_fin,
    df_cf,
    df_insider,
    df_snap,
    df_brokerage,
    df_tech,
    syms,
    cats,
)

if not summary.empty:
    grade_map = summary.set_index("symbol")[
        [c for c in ["Grade", "Performance Score", "Sales YoY %", "Net Profit YoY %"]
         if c in summary.columns]
    ].to_dict("index")
    cal_df["Grade"]             = cal_df["symbol"].map(lambda s: grade_map.get(s, {}).get("Grade", "—"))
    cal_df["Performance Score"] = cal_df["symbol"].map(lambda s: grade_map.get(s, {}).get("Performance Score"))
    cal_df["Sales YoY %"]       = cal_df["symbol"].map(lambda s: grade_map.get(s, {}).get("Sales YoY %"))
    cal_df["Net Profit YoY %"]  = cal_df["symbol"].map(lambda s: grade_map.get(s, {}).get("Net Profit YoY %"))
else:
    cal_df["Grade"] = "—"
    cal_df["Performance Score"] = None

# ─── Filter ───────────────────────────────────────────────────────────────────

display_df = cal_df[cal_df["in_watchlist"]].copy() if show_mode == "Watchlist stocks only" else cal_df.copy()
display_df = display_df.sort_values("result_date")

# ─── KPI strip ───────────────────────────────────────────────────────────────

total_results      = len(display_df)
today_results      = int((display_df["result_date"] == today).sum())
tomorrow_results   = int((display_df["result_date"] == today + timedelta(days=1)).sum())
watchlist_count    = int(display_df["in_watchlist"].sum())
a_grade_count      = int(display_df.get("Grade", pd.Series()).astype(str).isin(["A+", "A"]).sum())

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Total results",        total_results)
k2.metric("Declaring today",      today_results)
k3.metric("Declaring tomorrow",   tomorrow_results)
k4.metric("In your watchlist",    watchlist_count)
k5.metric("A/A+ grade",           a_grade_count)

st.divider()

# ─── Day-by-day grouped view ─────────────────────────────────────────────────

GRADE_COLOR = {
    "A+": "#166534", "A": "#16a34a", "B": "#ca8a04",
    "C": "#ea580c",  "F": "#dc2626", "—": "#6b7280",
}
GRADE_BG = {
    "A+": "#dcfce7", "A": "#d1fae5", "B": "#fef9c3",
    "C": "#ffedd5",  "F": "#fee2e2", "—": "#f3f4f6",
}

all_dates = sorted(display_df["result_date"].unique())

if not all_dates:
    st.info("No results for the current filter. Try switching to 'All NSE stocks'.")
    st.stop()

for result_date in all_dates:
    day_df = display_df[display_df["result_date"] == result_date].copy()
    day_df = day_df.sort_values(
        ["in_watchlist", "Performance Score"],
        ascending=[False, False],
    )

    delta = (result_date - today).days
    if   delta == 0:  day_label = "🔴 TODAY"
    elif delta == 1:  day_label = "🟠 TOMORROW"
    elif delta == -1: day_label = "✅ YESTERDAY"
    elif delta < 0:   day_label = f"✅ {abs(delta)} days ago"
    elif delta <= 3:  day_label = f"🟡 In {delta} days"
    else:             day_label = f"🔵 In {delta} days"

    wl_count = int(day_df["in_watchlist"].sum())
    header   = (
        f"{day_label} — **{result_date.strftime('%A, %d %b %Y')}** "
        f"&nbsp;·&nbsp; {len(day_df)} companies"
        + (f" &nbsp;·&nbsp; ⭐ {wl_count} in watchlist" if wl_count else "")
    )

    expanded = delta in (0, 1, -1)
    with st.expander(header, expanded=expanded):
        # ── Watchlist stocks first — card grid ────────────────────────────────
        wl_stocks = day_df[day_df["in_watchlist"]]
        other_stocks = day_df[~day_df["in_watchlist"]]

        if not wl_stocks.empty:
            st.markdown("**⭐ Your watchlist**")
            card_cols = st.columns(min(4, len(wl_stocks)))
            for idx, (_, row) in enumerate(wl_stocks.iterrows()):
                grade = str(row.get("Grade", "—"))
                g_col = GRADE_COLOR.get(grade, "#6b7280")
                g_bg  = GRADE_BG.get(grade,  "#f3f4f6")
                score = row.get("Performance Score")
                s_yoy = row.get("Sales YoY %")
                np_yoy = row.get("Net Profit YoY %")

                with card_cols[idx % 4]:
                    st.markdown(f"""
                    <div style="
                        border: 2px solid {g_col};
                        border-radius: 10px;
                        padding: 12px 14px;
                        background: {g_bg};
                        margin-bottom: 8px;
                    ">
                        <div style="display:flex; justify-content:space-between; align-items:center;">
                            <span style="font-weight:700; font-size:15px; color:#1e293b;">{row['symbol']}</span>
                            <span style="
                                background:{g_col}; color:white;
                                padding:2px 8px; border-radius:5px;
                                font-size:12px; font-weight:700;
                            ">{grade}</span>
                        </div>
                        <div style="font-size:12px; color:#475569; margin-top:3px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">
                            {row['company'][:30]}
                        </div>
                        <div style="font-size:12px; margin-top:6px; color:#334155;">
                            📊 Score: <b>{f"{score:.0f}" if pd.notna(score) else "—"}</b>
                            &nbsp;|&nbsp;
                            Sales: <b style="color:{'#16a34a' if pd.notna(s_yoy) and s_yoy > 0 else '#dc2626'}">
                                {f"{s_yoy:+.1f}%" if pd.notna(s_yoy) else "—"}
                            </b>
                        </div>
                        <div style="font-size:11px; margin-top:3px; color:#334155;">
                            NP: <b style="color:{'#16a34a' if pd.notna(np_yoy) and np_yoy > 0 else '#dc2626'}">
                                {f"{np_yoy:+.1f}%" if pd.notna(np_yoy) else "—"}
                            </b>
                            &nbsp;|&nbsp; {row['event']}
                        </div>
                        <a href="{row['tl_url']}" target="_blank"
                           style="font-size:11px; color:{g_col}; text-decoration:none;">
                           🔗 Trendlyne →
                        </a>
                    </div>
                    """, unsafe_allow_html=True)

        # ── All other stocks — compact table ──────────────────────────────────
        if not other_stocks.empty:
            if not wl_stocks.empty:
                st.markdown("**Other companies**")

            show_cols = [c for c in [
                "symbol", "company", "Grade", "Performance Score",
                "Sales YoY %", "Net Profit YoY %", "event",
            ] if c in other_stocks.columns]

            st.dataframe(
                other_stocks[show_cols].reset_index(drop=True).style.format(
                    {
                        "Performance Score": "{:.0f}",
                        "Sales YoY %":       "{:+.1f}%",
                        "Net Profit YoY %":  "{:+.1f}%",
                    },
                    na_rep="—",
                ),
                hide_index=True,
                width="stretch",
                height=min(420, 45 + len(other_stocks) * 36),
            )

st.divider()

# ─── Full week download ───────────────────────────────────────────────────────

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def build_calendar_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results Calendar"

    GRADE_HEX = {
        "A+": "166534", "A": "16a34a", "B": "ca8a04",
        "C":  "ea580c", "F": "dc2626",
    }

    headers = [
        "Date", "Day", "Symbol", "Company", "Grade",
        "Performance Score", "Sales YoY %", "Net Profit YoY %",
        "Event", "In Watchlist", "Trendlyne Link",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="1e3a5f")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        grade  = str(row.get("Grade", "—"))
        row_bg = "EFF6FF" if row.get("in_watchlist") else ("F8FAFC" if ri % 2 == 0 else "FFFFFF")

        values = [
            row["result_date"].strftime("%d-%b-%Y"),
            row["result_date"].strftime("%A"),
            row["symbol"],
            row["company"],
            grade,
            round(row["Performance Score"], 0) if pd.notna(row.get("Performance Score")) else "",
            round(row["Sales YoY %"],       1) if pd.notna(row.get("Sales YoY %"))       else "",
            round(row["Net Profit YoY %"],  1) if pd.notna(row.get("Net Profit YoY %"))  else "",
            row["event"],
            "✓" if row.get("in_watchlist") else "",
            row.get("tl_url", ""),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 5 and grade in GRADE_HEX:        # Grade
                cell.fill = PatternFill("solid", fgColor=GRADE_HEX[grade])
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            elif ci == 10 and val == "✓":              # Watchlist
                cell.font = Font(bold=True, color="166534")
                cell.fill = PatternFill("solid", fgColor="dcfce7")
            else:
                cell.fill = PatternFill("solid", fgColor=row_bg)

    col_widths = [13, 12, 14, 32, 8, 12, 12, 16, 20, 12, 40]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


dl1, dl2 = st.columns(2)
with dl1:
    st.download_button(
        "📥 Download full calendar (Excel)",
        data=build_calendar_excel(display_df),
        file_name=f"results_calendar_{start_date}_{end_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with dl2:
    # TradingView-ready watchlist of stocks declaring results this week
    wl_declaring = display_df[display_df["in_watchlist"]]["symbol"].unique()
    tv_string    = ",".join(f"NSE:{s}" for s in wl_declaring)
    st.download_button(
        "📺 TradingView: declaring-results watchlist (txt)",
        data=tv_string,
        file_name=f"tv_results_{start_date}_{end_date}.txt",
        mime="text/plain",
        use_container_width=True,
    )